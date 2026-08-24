import sys
import json
import os
import openpyxl


def read_xlsx(path, sheet=None, **_):
    wb = openpyxl.load_workbook(path, data_only=True)
    names = [sheet] if sheet else wb.sheetnames
    sheets = {}
    for name in names:
        if name not in wb.sheetnames:
            raise RuntimeError(f"找不到工作表: {name}（可用: {', '.join(wb.sheetnames)}）")
        ws = wb[name]
        sheets[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    return {"sheets": sheets, "sheetNames": wb.sheetnames}


def create_xlsx(path, sheets=None, overwrite=False, **_):
    if os.path.exists(path) and not overwrite:
        raise RuntimeError(f"檔案已存在，未帶 overwrite:true 不會覆蓋: {path}")
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    first = True
    for item in sheets or []:
        name = item.get("name", "Sheet")
        rows = item.get("rows", [])
        ws = default_sheet if first else wb.create_sheet(title=name)
        if first:
            ws.title = name
        first = False
        for row in rows:
            ws.append(row)
    wb.save(path)
    return {"path": path}


def append_row(path, sheet, values, **_):
    if not os.path.exists(path):
        raise RuntimeError(f"檔案不存在，請先用 create_xlsx 建立: {path}")
    wb = openpyxl.load_workbook(path)
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"找不到工作表: {sheet}（可用: {', '.join(wb.sheetnames)}）")
    ws = wb[sheet]
    ws.append(values)
    wb.save(path)
    return {"path": path, "row": ws.max_row}


def set_cell(path, sheet, row, col, value, **_):
    if not os.path.exists(path):
        raise RuntimeError(f"檔案不存在，請先用 create_xlsx 建立: {path}")
    wb = openpyxl.load_workbook(path)
    if sheet not in wb.sheetnames:
        raise RuntimeError(f"找不到工作表: {sheet}（可用: {', '.join(wb.sheetnames)}）")
    ws = wb[sheet]
    ws.cell(row=row, column=col, value=value)
    wb.save(path)
    return {"path": path, "row": row, "col": col}


def create_sheet(path, name, rows=None, **_):
    if not os.path.exists(path):
        raise RuntimeError(f"檔案不存在，請先用 create_xlsx 建立: {path}")
    wb = openpyxl.load_workbook(path)
    if name in wb.sheetnames:
        raise RuntimeError(f"工作表已存在: {name}")
    ws = wb.create_sheet(title=name)
    for row in rows or []:
        ws.append(row)
    wb.save(path)
    return {"path": path, "sheet": name}


def extract_images(path, output_dir, **_):
    if not os.path.exists(path):
        raise RuntimeError(f"檔案不存在: {path}")
    os.makedirs(output_dir, exist_ok=True)
    wb = openpyxl.load_workbook(path)
    images = []
    for ws in wb.worksheets:
        for img in getattr(ws, "_images", []):
            anchor = getattr(img, "anchor", None)
            cell_from = getattr(anchor, "_from", None) if anchor is not None else None
            row = (cell_from.row + 1) if cell_from is not None else None
            col = (cell_from.col + 1) if cell_from is not None else None
            data = img._data()
            src_name = getattr(img, "path", None) or getattr(img, "filename", None) or ""
            ext = os.path.splitext(src_name)[1] or ".png"
            seq = len(images) + 1
            out_path = os.path.join(output_dir, f"image_{seq:03d}{ext}")
            with open(out_path, "wb") as f:
                f.write(data)
            images.append({"file": out_path, "sheet": ws.title, "anchor_row": row, "anchor_col": col})
    return {"images": images, "count": len(images)}


ACTIONS = {
    "read": read_xlsx,
    "create": create_xlsx,
    "append_row": append_row,
    "set_cell": set_cell,
    "create_sheet": create_sheet,
    "extract_images": extract_images,
}


def main():
    payload = json.loads(sys.stdin.buffer.read().decode("utf-8"))
    action = payload.get("action")
    if action not in ACTIONS:
        raise RuntimeError(f"未知的 action: {action}（可用: {', '.join(ACTIONS)}）")
    result = ACTIONS[action](**{k: v for k, v in payload.items() if k != "action"})
    print(json.dumps({"success": True, **result}, ensure_ascii=False))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    try:
        main()
    except Exception as e:
        print(json.dumps({"success": False, "message": str(e)}, ensure_ascii=False))
        sys.exit(1)
